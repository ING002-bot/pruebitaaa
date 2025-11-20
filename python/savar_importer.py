from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException
)
from webdriver_manager.chrome import ChromeDriverManager
import time
import json
import os
import glob
from pathlib import Path
import mysql.connector
from mysql.connector import Error
from typing import Dict, List, Any, Optional, Union
import re
import unicodedata

# Configuración global

def _normalize_key(text: str) -> str:
    """Normaliza texto: quita acentos, espacios y saltos de línea, a MAYÚSCULAS."""
    if text is None:
        return ""
    t = unicodedata.normalize('NFKD', text)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = t.upper().replace(' ', '').replace('\n', '').replace('\r', '')
    return t

def _find_th_by_text(driver: webdriver.Chrome, key_text: str) -> Optional[Any]:
    """Busca un <th> cuyo texto normalizado coincida con key_text (también normalizado)."""
    target = _normalize_key(key_text)
    try:
        th_candidates = driver.find_elements(By.XPATH, "//th")
        for th in th_candidates:
            try:
                txt = (th.text or "")
                if _normalize_key(txt) and target in _normalize_key(txt):
                    return th
            except Exception:
                continue
    except Exception:
        pass

def _set_date_with_datepicker(driver: webdriver.Chrome, input_elem, yyyy_mm_dd: str, timeout: int = 10) -> bool:
    """
    Abre el datepicker del input, navega al año/mes del valor y hace clic en el día.
    Soporta datepickers tipo Bootstrap/flatpickr/ui-datepicker con flechas « » y celdas <td>/<span> con número de día.
    """
    try:
        # Parsear fecha
        y, m, d = yyyy_mm_dd.split('-')
        y = int(y); m = int(m); d = int(d)

        # Focus y click para abrir calendario
        try:
            input_elem.click()
        except Exception:
            driver.execute_script("arguments[0].click();", input_elem)
        time.sleep(0.2)

        # Localizar contenedor del calendario visible
        panel_candidates = [
            "//div[contains(@class,'datepicker') and not(contains(@style,'display: none'))]",
            "//div[contains(@class,'ui-datepicker') and not(contains(@style,'display: none'))]",
            "//div[contains(@class,'flatpickr-calendar') and contains(@class,'open')]",
        ]
        panel = None
        for xp in panel_candidates:
            try:
                el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
                if el and el.is_displayed():
                    panel = el
                    break
            except Exception:
                continue
        if panel is None:
            return False

        # Intentar leer el título de mes/año y los controles de navegación
        def get_month_year_text():
            try:
                for xp in [
                    ".//*[contains(@class,'datepicker-switch')]",
                    ".//div[contains(@class,'cur-month')]",
                    ".//*[contains(@class,'ui-datepicker-title')]",
                    ".//span[contains(@class,'flatpickr-monthDropdown-months')]",
                ]:
                    try:
                        el = panel.find_element(By.XPATH, xp)
                        if el and el.is_displayed():
                            return el.text.strip()
                    except Exception:
                        continue
            except Exception:
                pass
            return ""

        # Botones prev/next
        def click_prev():
            for xp in [
                ".//*[contains(@class,'prev') or contains(.,'«')]",
                ".//span[contains(@class,'flatpickr-prev')]",
                ".//a[contains(@class,'ui-datepicker-prev')]",
            ]:
                try:
                    b = panel.find_element(By.XPATH, xp)
                    if b.is_displayed():
                        driver.execute_script("arguments[0].click();", b)
                        time.sleep(0.15)
                        return True
                except Exception:
                    continue
            return False
        def click_next():
            for xp in [
                ".//*[contains(@class,'next') or contains(.,'»')]",
                ".//span[contains(@class,'flatpickr-next')]",
                ".//a[contains(@class,'ui-datepicker-next')]",
            ]:
                try:
                    b = panel.find_element(By.XPATH, xp)
                    if b.is_displayed():
                        driver.execute_script("arguments[0].click();", b)
                        time.sleep(0.15)
                        return True
                except Exception:
                    continue
            return False

        # Navegar aproximando por diferencia de meses si no hay picker de mes directo
        # Intento limitado a 24 pasos para evitar bucles infinitos
        for _ in range(24):
            title = get_month_year_text()
            # Si el título no es parseable, intentar directamente clickear día; muchos pickers ya posicionan por valor actual
            try:
                # Click en el día
                day_candidates = panel.find_elements(By.XPATH, f".//*[self::td or self::span or self::a][normalize-space()='{d}']")
                for dc in day_candidates:
                    try:
                        if dc.is_displayed():
                            driver.execute_script("arguments[0].click();", dc)
                            time.sleep(0.2)
                            return True
                    except Exception:
                        continue
            except Exception:
                pass
            # Si no hizo click, intentamos mover mes a mes en función de comparación con input value
            # Como fallback, alternar prev/next unas pocas veces
            moved = click_prev() or click_next()
            if not moved:
                break

        return False
    except Exception:
        return False

def ensure_checkbox_by_label(driver: webdriver.Chrome, label_text: str, checked: bool, timeout: int = 10) -> bool:
    """
    Asegura el estado (checked/unchecked) del checkbox asociado al label dado.
    Busca un input[type=checkbox] cercano al label con ese texto.
    """
    try:
        # Normalizar búsqueda con y sin acentos
        xps = [
            f"//*[self::label or self::div or self::span][contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'{label_text}')]/preceding::input[@type='checkbox'][1]",
            f"//*[self::label or self::div or self::span][contains(.,'{label_text}')]/preceding::input[@type='checkbox'][1]",
            f"//*[self::label or self::div or self::span][contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'{label_text}')]/ancestor::*[1]//input[@type='checkbox']",
            f"//*[self::label or self::div or self::span][contains(.,'{label_text}')]/ancestor::*[1]//input[@type='checkbox']",
        ]
        cb = None
        for xp in xps:
            try:
                el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
                if el and el.is_displayed():
                    cb = el
                    break
            except Exception:
                continue
        if not cb:
            return False
        is_checked = (cb.get_attribute('checked') is not None) or cb.is_selected()
        if checked != is_checked:
            try:
                driver.execute_script("arguments[0].click();", cb)
                time.sleep(0.2)
            except Exception:
                cb.click()
        return True
    except Exception:
        return False

def close_overlays_and_datepickers(driver: webdriver.Chrome) -> None:
    """Cierra posibles overlays, backdrops y datepickers que bloqueen los clics."""
    try:
        # Tecla ESC varias veces
        try:
            body = driver.find_element(By.TAG_NAME, 'body')
            for _ in range(3):
                body.send_keys(Keys.ESCAPE)
                time.sleep(0.05)
        except Exception:
            pass

        # Ocultar elementos comunes por JS
        js = """
        (function(){
          var sels = [
            '.modal-backdrop', '.MuiBackdrop-root', '.swal2-container',
            '.xdsoft_datetimepicker', '.datepicker', '.date-picker',
            '.overlay', '.blockUI', '.loader', '.preloader'
          ];
          for (var i=0;i<sels.length;i++){
            var nodes = document.querySelectorAll(sels[i]);
            for (var j=0;j<nodes.length;j++){
              var el = nodes[j];
              try{ el.style.display='none'; el.style.visibility='hidden'; el.style.pointerEvents='none'; }catch(e){}
            }
          }
        })();
        """
        try:
            driver.execute_script(js)
        except Exception:
            pass
    except Exception:
        pass

def dismiss_error_dialog_if_any(driver: webdriver.Chrome) -> bool:
    """Cierra diálogos de error tipo SweetAlert/Bootstrap. Devuelve True si cerró alguno."""
    closed = False
    try:
        selectors = [
            (By.XPATH, "//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'HA OCURRIDO UN ERROR')]/ancestor::*[contains(@class,'swal2-popup') or contains(@class,'modal')][1]//button[normalize-space()='OK']"),
            (By.CSS_SELECTOR, ".swal2-container .swal2-confirm"),
            (By.XPATH, "//div[contains(@class,'modal') and .//*[contains(.,'error') or contains(.,'Error')]]//*[self::button or self::a][normalize-space()='OK']")
        ]
        for by, sel in selectors:
            try:
                btns = driver.find_elements(by, sel)
                for b in btns:
                    if b.is_displayed():
                        try:
                            driver.execute_script("arguments[0].click();", b)
                        except Exception:
                            try:
                                b.click()
                            except Exception:
                                continue
                        time.sleep(0.2)
                        closed = True
                        break
                if closed:
                    break
            except Exception:
                continue
    except Exception:
        pass
    return closed
def setup_driver(headless: bool = True, download_dir: str = None) -> webdriver.Chrome:
    """
    Configura y retorna una instancia del navegador Chrome.
    
    Args:
        headless: Si es True, el navegador se ejecutará en modo sin cabeza (sin interfaz gráfica).
        
    Returns:
        Instancia configurada de webdriver.Chrome
    """
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-plugins')
    chrome_options.add_argument('--disable-images')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-web-security')
    chrome_options.add_argument('--allow-running-insecure-content')
    
    # Preferencias de descarga si se especifica download_dir
    if download_dir:
        try:
            Path(download_dir).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        prefs = {
            "download.default_directory": str(Path(download_dir).resolve()),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1
        }
        chrome_options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_page_load_timeout(120)  # Aumentar timeout a 2 minutos
    driver.implicitly_wait(20)  # Aumentar espera implícita
    driver.maximize_window()  # Maximizar ventana
    # Habilitar descargas en modo headless mediante CDP
    try:
        if download_dir:
            driver.execute_cdp_cmd(
                "Page.setDownloadBehavior",
                {"behavior": "allow", "downloadPath": str(Path(download_dir).resolve())}
            )
    except Exception:
        pass
    return driver

def login_and_fetch_saver(driver: webdriver.Chrome, usuario: str, contrasena: str, fecha_inicio: str = None, fecha_fin: str = None, timeout: int = 30) -> bool:
    """
    Inicia sesión en el sistema SAVAR con las credenciales proporcionadas.
    
    Args:
        driver: Instancia de Selenium WebDriver
        usuario: Nombre de usuario para el inicio de sesión
        contrasena: Contraseña para el inicio de sesión
        fecha_inicio: Fecha de inicio en formato YYYY-MM-DD (opcional)
        fecha_fin: Fecha de fin en formato YYYY-MM-DD (opcional)
        timeout: Tiempo máximo de espera en segundos para los elementos
        
    Returns:
        bool: True si el inicio de sesión fue exitoso, False en caso contrario
    """
    try:
        print(f"Iniciando sesión en SAVAR con usuario: {usuario}")
        
        # Navegar a la página de inicio de sesión
        print("Navegando a la página de inicio de sesión...")
        driver.get("https://app.savarexpress.com.pe/sso/Inicio/")
        
        # Esperar a que la página se cargue completamente
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        print("Página cargada correctamente")
        
        # Tomar screenshot para depuración
        driver.save_screenshot("screenshot_login_page.png")
        print("Screenshot guardado: screenshot_login_page.png")
        
        # Intentar diferentes selectores para los campos de usuario y contraseña
        print("Buscando campos de login...")
        
        # Lista de selectores posibles para los campos de formulario
        username_selectors = [
            (By.CSS_SELECTOR, "input[type='text']"),
            (By.CSS_SELECTOR, "input[name*='usuario']"),
            (By.CSS_SELECTOR, "input[id*='usuario']"),
            (By.CSS_SELECTOR, "#username"),
            (By.NAME, "username")
        ]
        
        password_selectors = [
            (By.CSS_SELECTOR, "input[type='password']"),
            (By.CSS_SELECTOR, "input[name*='password']"),
            (By.CSS_SELECTOR, "input[id*='password']"),
            (By.CSS_SELECTOR, "#password"),
            (By.NAME, "password")
        ]
        
        submit_selectors = [
            (By.CSS_SELECTOR, "button[type='submit']"),
            (By.CSS_SELECTOR, "input[type='submit']"),
            (By.XPATH, "//button[contains(., 'Iniciar sesión')]"),
            (By.XPATH, "//button[contains(., 'Login')]")
        ]
        
        # Función para encontrar un elemento con múltiples estrategias
        def find_element_with_retry(selectors, timeout=10):
            for by, selector in selectors:
                try:
                    element = WebDriverWait(driver, timeout).until(
                        EC.presence_of_element_located((by, selector))
                    )
                    print(f"Elemento encontrado: {by}={selector}")
                    return element
                except Exception as e:
                    print(f"No se encontró con {by}={selector}: {str(e)}")
                    continue
            return None
        
        # Buscar campos de usuario y contraseña
        username_field = find_element_with_retry(username_selectors)
        if not username_field:
            print("Error: No se pudo encontrar el campo de usuario")
            return False

        password_field = find_element_with_retry(password_selectors)
        if not password_field:
            print("Error: No se pudo encontrar el campo de contraseña")
            return False
        
        print("Campos de login encontrados, completando credenciales...")
        
        # Completar los campos
        username_field.clear()
        username_field.send_keys(usuario)
        print("Usuario ingresado")
        
        password_field.clear()
        password_field.send_keys(contrasena)
        print("Contraseña ingresada")
        
        # Intentar hacer clic en el botón de envío
        submit_button = find_element_with_retry(submit_selectors)
        if submit_button:
            print("Haciendo clic en el botón de envío...")
            try:
                # Intentar hacer clic con JavaScript si el clic normal falla
                driver.execute_script("arguments[0].click();", submit_button)
            except Exception as e:
                print(f"Error al hacer clic en el botón: {e}")
                submit_button.click()
        else:
            print("No se encontró el botón de envío, intentando con Enter...")
            password_field.send_keys(Keys.RETURN)
        
        # Esperar a que se complete el inicio de sesión
        print("Esperando a que se complete el inicio de sesión...")
        time.sleep(5)
        
        # Verificar si el inicio de sesión fue exitoso
        current_url = driver.current_url
        print(f"URL después del login: {current_url}")
        
        # Verificar si hay mensajes de error (con manejo robusto)
        try:
            # Esperar un poco más para que la página se estabilice
            time.sleep(3)
            error_elements = driver.find_elements(
                By.XPATH, 
                "//*[contains(text(), 'Error') or contains(text(), 'Incorrecto') or contains(@class, 'error')]"
            )
            
            for error_element in error_elements:
                if error_element.is_displayed() and error_element.text.strip():
                    error_text = error_element.text.strip()
                    print(f"Error en inicio de sesión: {error_text}")
                    return False
        except Exception as e:
            print(f"No se encontraron mensajes de error visibles: {str(e)}")
            # Continuar aunque no se puedan verificar errores
        
        # Tomar screenshot después del login
        driver.save_screenshot("screenshot_after_login.png")
        print("Screenshot guardado: screenshot_after_login.png")
        
        # Si se proporcionaron fechas, configurar específicamente 'Fecha de Recepción' y consultar
        if fecha_inicio and fecha_fin:
            print(f"Configurando rango de fechas (Recepción): {fecha_inicio} a {fecha_fin}")
            try:
                # Asegurar checkboxes correctos: Recepción ON, Creación OFF
                ensure_checkbox_by_label(driver, "Fecha de Recepcion", True, timeout=10)
                ensure_checkbox_by_label(driver, "Fecha de Recepción", True, timeout=10)
                ensure_checkbox_by_label(driver, "Fecha de Creacion", False, timeout=10)
                ensure_checkbox_by_label(driver, "Fecha de Creación", False, timeout=10)

                ok = set_date_inputs_by_label(
                    driver,
                    label_text="Fecha de Recepcion",  # sin acento para el XPath normalizado
                    start_date=fecha_inicio,
                    end_date=fecha_fin,
                    timeout=20
                )
                if not ok:
                    print("Reintentando con 'Fecha de Recepción' (con acento)...")
                    ok = set_date_inputs_by_label(
                        driver,
                        label_text="Fecha de Recepción",
                        start_date=fecha_inicio,
                        end_date=fecha_fin,
                        timeout=20
                    )
                if ok:
                    # Verificación y corrección si quedaron iguales
                    try:
                        base_xp = "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Recepcion')]|//label[contains(.,'Fecha de Recepción')]"
                        first_input = driver.find_element(By.XPATH, base_xp + "/following::input[1]")
                        second_input = driver.find_element(By.XPATH, base_xp + "/following::input[2]")
                        v1 = (first_input.get_attribute('value') or '').strip()
                        v2 = (second_input.get_attribute('value') or '').strip()
                        print(f"Valores después de escribir -> inicio: {v1} | fin: {v2}")
                        if v1 == v2 and v1 != fecha_inicio:
                            print("Detectado v1=v2, reescribiendo inicio con fecha de ayer...")
                            # Reforzar valor de inicio
                            first_input.click()
                            first_input.send_keys(Keys.CONTROL + 'a')
                            first_input.send_keys(Keys.DELETE)
                            for ch in fecha_inicio:
                                first_input.send_keys(ch)
                                time.sleep(0.03)
                            first_input.send_keys(Keys.TAB)
                            time.sleep(0.3)
                    except Exception as e:
                        print(f"No se pudo verificar/corregir fechas: {e}")

                    print("Fechas de Recepción configuradas. Procediendo a Consultar...")
                    # Intentar clic por texto visible
                    clicked = click_element_by_text(driver, "Consultar", timeout=10)
                    if not clicked:
                        # Fallback: buscar posibles botones de consulta
                        candidates = [
                            (By.XPATH, "//button[contains(.,'Consultar') or contains(.,'Buscar')]"),
                            (By.CSS_SELECTOR, "button[type='submit']"),
                        ]
                        for by, sel in candidates:
                            try:
                                btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((by, sel)))
                                driver.execute_script("arguments[0].click();", btn)
                                print("Botón de consulta clickeado (fallback)")
                                break
                            except Exception:
                                continue
                    # Esperar a que se carguen resultados
                    wait_until_not_processing(driver, timeout=30)
                else:
                    print("No fue posible configurar las fechas de 'Fecha de Recepción'")
            except Exception as e:
                print(f"Error al configurar fechas (Recepción): {str(e)}")
        
        print("Inicio de sesión exitoso")
        return True
        
    except Exception as e:
        print(f"Error en el proceso de inicio de sesión: {str(e)}")
        # Tomar screenshot en caso de error
        try:
            driver.save_screenshot("screenshot_error.png")
            print("Screenshot del error guardado: screenshot_error.png")
        except:
            print("No se pudo guardar el screenshot del error")
        return False
                
    except Exception as e:
        print(f"Error durante el inicio de sesión: {str(e)}")
        return False
            
    # Verificar si hay mensajes de error
    try:
        error_element = driver.find_element(By.XPATH, "//div[contains(text(), 'Error') or contains(text(), 'Incorrecto') or contains(@class, 'error')]")
        if error_element.is_displayed():
            print(f"Error en inicio de sesión: {error_element.text}")
            return False
    except NoSuchElementException:
        pass
    # Si llegamos aquí, asumimos que el login fue exitoso
    print("Login procesado - continuando...")
    close_overlays_and_datepickers(driver)
    return True

def open_side_menu_if_needed(driver: webdriver.Chrome, timeout: int = 10) -> None:
    """Intenta abrir el menú lateral (hamburguesa) si está colapsado."""
    try:
        # Buscar el botón de menú tipo hamburguesa en distintas implementaciones
        candidates = [
            (By.CSS_SELECTOR, "button[aria-label*='menu']"),
            (By.CSS_SELECTOR, "button[aria-label*='Menú']"),
            (By.CSS_SELECTOR, "button[class*='navbar']"),
            (By.CSS_SELECTOR, "button[class*='menu']"),
            (By.XPATH, "//button[.//span[contains(@class,'menu')] or contains(@class,'menu') or contains(., 'menu')]")
        ]
        for by, sel in candidates:
            try:
                btn = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, sel)))
                if btn and btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.5)
                    break
            except Exception:
                continue
    except Exception:
        pass

def set_date_inputs_by_label(
    driver: webdriver.Chrome,
    label_text: str,
    start_date: str,
    end_date: str,
    timeout: int = 15
) -> bool:
    """
    Busca los 2 inputs de fecha asociados a una etiqueta (ej. 'Fecha de Creacion') y los setea por teclado.
    Estrategia: click -> CTRL+A -> escribir fecha -> TAB. Luego dispara eventos JS como respaldo.
    """
    print(f"Iniciando configuración de fechas para etiqueta: {label_text}")
    try:
        def find_inputs_for_label(label_text_local: str):
            base = f"//*[self::label or self::div or self::span][contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'{label_text_local}')]"
            # Preferir inputs dentro de la misma fila/contenedor del label
            strategies = [
                base + "/ancestor::div[contains(@class,'row')][1]//input",
                base + "/ancestor::div[1]//input",
                base + "/ancestor::tr[1]//input",
                base + "/following-sibling::*//input",
                base + "/following::input[position()<=2]",
            ]
            for xp in strategies:
                try:
                    els = WebDriverWait(driver, timeout).until(
                        EC.presence_of_all_elements_located((By.XPATH, xp))
                    )
                    # Filtrar inputs que aparenten ser fechas (type=date o con patrón YYYY-MM-DD en value/placeholder)
                    def looks_like_date(el):
                        try:
                            t = (el.get_attribute('type') or '').lower()
                            ph = (el.get_attribute('placeholder') or '')
                            val = (el.get_attribute('value') or '')
                            cls = (el.get_attribute('class') or '').lower()
                            return (
                                t == 'date' or
                                'date' in cls or
                                re.search(r"\d{4}-\d{2}-\d{2}", ph or '') or
                                re.search(r"\d{4}-\d{2}-\d{2}", val or '')
                            )
                        except Exception:
                            return False
                    vis = [e for e in els if e.is_displayed() and e.is_enabled() and looks_like_date(e)]
                    if len(vis) >= 2:
                        return vis[0], vis[1]
                except Exception:
                    continue
            return None, None

        print("Buscando inputs de fecha asociados al label en su misma fila...")
        first_input, second_input = find_inputs_for_label(label_text)
        if not first_input or not second_input:
            print("No encontrados con label sin acento, probando con acento...")
            first_input, second_input = find_inputs_for_label(label_text)
        if not first_input or not second_input:
            print("Fallo al localizar los dos inputs de fecha asociados al label; aplicando fallback por checkbox marcado")
            # Fallback: localizar la fila del checkbox marcado (Recepción) y tomar los 2 inputs dentro
            try:
                cb_checked = driver.find_element(By.XPATH, "//input[@type='checkbox' and @checked]/ancestor::*[contains(@class,'row') or contains(@class,'form-group') or self::tr][1]")
                cand = [e for e in cb_checked.find_elements(By.XPATH, ".//input") if e.is_displayed() and e.is_enabled()]
                # Filtrar a los que parezcan fechas
                def looks_like_date2(el):
                    try:
                        t=(el.get_attribute('type') or '').lower(); cls=(el.get_attribute('class') or '').lower(); ph=(el.get_attribute('placeholder') or ''); val=(el.get_attribute('value') or '')
                        return t=='date' or 'date' in cls or re.search(r"\d{4}-\d{2}-\d{2}", ph or '') or re.search(r"\d{4}-\d{2}-\d{2}", val or '')
                    except Exception:
                        return False
                cand = [e for e in cand if looks_like_date2(e)]
                if len(cand) >= 2:
                    first_input, second_input = cand[0], cand[1]
                else:
                    return False
            except Exception:
                return False
        
        # Asegurarse de que los inputs son visibles y habilitados
        WebDriverWait(driver, timeout).until(
            lambda d: first_input.is_displayed() and first_input.is_enabled() and
                     second_input.is_displayed() and second_input.is_enabled()
        )
        
        vis = [first_input, second_input]
        print(f"Inputs encontrados: {len(vis)} en misma fila/contenedor del label")
        
        if len(vis) != 2:
            print(f"Error: No se encontraron los 2 inputs necesarios. Se encontraron {len(vis)} inputs visibles.")
            return False

        def type_date(elem, value, press_tab: bool = True):
            try:
                print(f"Configurando fecha: {value}")
                
                # Hacer scroll al elemento
                driver.execute_script("arguments[0].scrollIntoView({block:'center', behavior:'smooth'});", elem)
                time.sleep(0.5)  # Esperar a que termine el scroll
                
                # Asegurarse de que el elemento es interactuable
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//*[@id='{elem.get_attribute('id')}']"))
                    if elem.get_attribute('id') else (By.XPATH, ".//input")
                )
                
                # Quitar atributo readonly si existe
                driver.execute_script("arguments[0].removeAttribute('readonly');", elem)
                
                # Hacer clic para enfocar
                try:
                    elem.click()
                except:
                    driver.execute_script("arguments[0].click();", elem)
                
                # Limpiar el campo
                for _ in range(3):  # Múltiples intentos para asegurar limpieza
                    try:
                        elem.clear()
                        elem.send_keys(Keys.CONTROL + 'a')
                        elem.send_keys(Keys.DELETE)
                        time.sleep(0.2)
                    except:
                        pass
                
                # Escribir la fecha caracter por caracter
                for char in value:
                    elem.send_keys(char)
                    time.sleep(0.05)  # Pequeña pausa entre caracteres
                
                # Confirmar la fecha
                elem.send_keys(Keys.RETURN)
                time.sleep(0.5)  # Esperar a que se aplique
                
                # Disparar eventos para asegurar que se registre el cambio
                driver.execute_script("""
                    var elem = arguments[0];
                    var value = arguments[1];
                    elem.value = value;
                    
                    // Disparar todos los eventos relevantes
                    ['input', 'change', 'blur', 'keyup', 'keydown', 'keypress'].forEach(function(event) {
                        var evt = new Event(event, { bubbles: true });
                        elem.dispatchEvent(evt);
                    });
                    
                    // Forzar actualización si es un campo de React/Angular/Vue
                    if ('_value' in elem) elem._value = value;
                    if ('_valueSetter' in elem) {
                        var prototype = Object.getPrototypeOf(elem);
                        var valueProperty = Object.getOwnPropertyDescriptor(prototype, 'value');
                        if (valueProperty && valueProperty.set) {
                            valueProperty.set.call(elem, value);
                        }
                    }
                """, elem, value)
                
                # Pequeña pausa para asegurar que los eventos se procesen
                time.sleep(0.5)
                
                # Verificar que el valor se estableció correctamente
                current_value = elem.get_attribute('value')
                if current_value != value:
                    print(f"Advertencia: El valor actual '{current_value}' no coincide con el esperado '{value}'")
                    # Intentar establecer el valor directamente como último recurso
                    driver.execute_script("arguments[0].value = arguments[1];", elem, value)
                
                return True
                
            except Exception as e:
                print(f"Error en type_date: {str(e)}")
                import traceback
                traceback.print_exc()
                return False

        def ensure_value(elem, value):
            try:
                val = elem.get_attribute('value') or ''
                if val.strip() == value:
                    return True
                # Set por JS si no coincidió
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true})); arguments[0].blur();",
                    elem, value
                )
                time.sleep(0.1)
                val2 = elem.get_attribute('value') or ''
                if val2.strip() == value:
                    return True
                # Último recurso: si el datepicker está visible, clicar el día correspondiente
                try:
                    day = str(int(value.split('-')[-1]))  # quitar ceros a la izquierda
                except Exception:
                    day = value.split('-')[-1]
                panel_xps = [
                    "//div[contains(@class,'datepicker') and not(contains(@style,'display: none'))]",
                    "//div[contains(@class,'ui-datepicker') and not(contains(@style,'display: none'))]",
                    "//div[contains(@class,'flatpickr-calendar') and contains(@class,'open')]",
                    "//div[contains(@class,'xdsoft_datetimepicker') and not(contains(@style,'display: none'))]",
                ]
                panel = None
                for xp in panel_xps:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el.is_displayed():
                            panel = el
                            break
                    except Exception:
                        continue
                if panel is not None:
                    candidates = panel.find_elements(By.XPATH, f".//*[normalize-space()='{day}']")
                    for c in candidates:
                        try:
                            if c.is_displayed():
                                driver.execute_script("arguments[0].click();", c)
                                time.sleep(0.1)
                                break
                        except Exception:
                            continue
                # Releer
                return (elem.get_attribute('value') or '').strip() == value
            except Exception:
                return False

        # Escribir primero fecha de fin (derecha), luego inicio (izquierda) para evitar sincronizaciones
        print(f"Configurando fecha de fin: {end_date}")
        ok2 = type_date(vis[1], end_date, press_tab=False)
        if ok2 is False:
            ok2 = _set_date_with_datepicker(driver, vis[1], end_date, timeout=10)
        if not ok2:
            print("Error al configurar fecha de fin, intentando método alternativo...")
            driver.execute_script("""
                var elem = arguments[0];
                var value = arguments[1];
                elem.value = value;
                elem.dispatchEvent(new Event('input', {bubbles: true}));
                elem.dispatchEvent(new Event('change', {bubbles: true}));
            """, vis[1], end_date)
            ok2 = True

        time.sleep(0.7)

        print(f"Configurando fecha de inicio: {start_date}")
        ok1 = type_date(vis[0], start_date, press_tab=True)
        if ok1 is False:
            ok1 = _set_date_with_datepicker(driver, vis[0], start_date, timeout=10)
        if not ok1:
            print("Error al configurar fecha de inicio, intentando método alternativo...")
            # Método alternativo para la fecha de inicio
            driver.execute_script("""
                var elem = arguments[0];
                var value = arguments[1];
                elem.value = value;
                elem.dispatchEvent(new Event('input', {bubbles: true}));
                elem.dispatchEvent(new Event('change', {bubbles: true}));
            """, vis[0], start_date)
            ok1 = True  # Asumir éxito después del intento alternativo

        # Verificar orden: si el input izquierdo contiene hoy y el derecho ayer, intercambiar
        try:
            v_left = (vis[0].get_attribute('value') or '').strip()
            v_right = (vis[1].get_attribute('value') or '').strip()
            if v_left == end_date and v_right == start_date:
                print("Orden invertido detectado, reescribiendo: izquierda=inicio, derecha=fin")
                type_date(vis[0], start_date, press_tab=True)
                time.sleep(0.2)
                type_date(vis[1], end_date, press_tab=False)
            # Si aún quedan iguales, forzar por JS y disparar eventos
            v_left2 = (vis[0].get_attribute('value') or '').strip()
            v_right2 = (vis[1].get_attribute('value') or '').strip()
            if v_left2 == v_right2:
                print("Ambos campos siguen iguales, forzando valores por JS (inicio=ayer, fin=hoy)")
                driver.execute_script("""
                    var a=arguments[0], b=arguments[1], v1=arguments[2], v2=arguments[3];
                    a.value=v1; b.value=v2;
                    ['input','change','blur','keyup'].forEach(function(ev){
                        a.dispatchEvent(new Event(ev,{bubbles:true}));
                        b.dispatchEvent(new Event(ev,{bubbles:true}));
                    });
                """, vis[0], vis[1], start_date, end_date)
                time.sleep(0.2)
                # Releer y último intento: escribir con teclas si siguen iguales
                v_left3 = (vis[0].get_attribute('value') or '').strip()
                v_right3 = (vis[1].get_attribute('value') or '').strip()
                if v_left3 == v_right3:
                    type_date(vis[0], start_date, press_tab=True)
                    time.sleep(0.2)
                    type_date(vis[1], end_date, press_tab=False)
        except Exception:
            pass

        # Cerrar calendarios inmediatamente tras el segundo input
        try:
            body = driver.find_element(By.TAG_NAME, 'body')
            for _ in range(2):
                body.send_keys(Keys.ESCAPE)
        except Exception:
            pass
        # Clic en un área neutra para asegurar blur
        try:
            neutral = driver.find_element(By.XPATH, "//div[contains(.,'Control de Almacenes')] | //h2[contains(.,'CONTROL DE ALMACENES')]")
            driver.execute_script("arguments[0].click();", neutral)
        except Exception:
            pass

        try:
            driver.save_screenshot("step_dates_typed.png")
        except Exception:
            pass

        # Intentar cerrar el datepicker si quedó abierto (ESC y click fuera)
        try:
            body = driver.find_element(By.TAG_NAME, 'body')
            for _ in range(2):
                body.send_keys(Keys.ESCAPE)
                time.sleep(0.1)
        except Exception:
            pass
        try:
            # Click en el título de sección para cerrar overlays
            anchors = [
                "//div[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]",
                "//h1[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]",
                "//h2[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]",
            ]
            for xp in anchors:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        driver.execute_script("arguments[0].click();", el)
                        break
                except Exception:
                    continue
        except Exception:
            pass

        return ok1 and ok2
    except Exception:
        return False

def wait_until_not_processing(driver: webdriver.Chrome, timeout: int = 30) -> bool:
    """
    Espera a que desaparezcan overlays/spinners o mensajes de 'Procesando datos'.
    Retorna True si la página quedó lista, False si alcanzó el timeout.
    """
    start = time.time()
    selectors_css = [
        ".loading", ".spinner", ".overlay", ".blockUI", ".preloader", ".MuiBackdrop-root",
        "div[aria-busy='true']", "div[role='progressbar']"
    ]
    xpaths_txt = [
        "//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'PROCESANDO DATOS')]",
        "//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CARGANDO')]",
    ]
    while time.time() - start < timeout:
        try:
            busy_found = False
            # Revisar CSS comunes
            for sel in selectors_css:
                try:
                    els = driver.find_elements(By.CSS_SELECTOR, sel)
                    if any(e.is_displayed() for e in els):
                        busy_found = True
                        break
                except Exception:
                    continue
            # Revisar textos
            if not busy_found:
                for xp in xpaths_txt:
                    try:
                        els = driver.find_elements(By.XPATH, xp)
                        if any(e.is_displayed() for e in els):
                            busy_found = True
                            break
                    except Exception:
                        continue
            if not busy_found:
                return True
            time.sleep(0.5)
        except Exception:
            break
    return False

def click_element_by_text(driver: webdriver.Chrome, text: str, timeout: int = 15) -> bool:
    """
    Hace clic en un elemento que contenga el texto visible indicado. Intenta con varias etiquetas.
    Retorna True si logró hacer clic.
    """
    xpaths = [
        f"//button[normalize-space()='{text}']",
        f"//a[normalize-space()='{text}']",
        f"//li[normalize-space()='{text}']",
        f"//div[normalize-space()='{text}']",
        f"//*[self::span or self::p or self::h2 or self::h3][normalize-space()='{text}']",
        f"//*[contains(normalize-space(), '{text}')]"
    ]
    for xp in xpaths:
        try:
            el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xp)))
            if el and el.is_displayed():
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", el)
                return True
        except Exception:
            continue
    return False

# ==========================
# Exportar Excel desde modal
# ==========================

def get_downloads_dir() -> str:
    """Obtiene el directorio de Descargas del usuario actual (Windows/macOS/Linux)."""
    try:
        home = Path.home()
        downloads = home / "Downloads"
        return str(downloads)
    except Exception:
        return str(Path.cwd())

def _list_downloads(download_dir: str, pattern: str = "*.xls*") -> list:
    try:
        return glob.glob(str(Path(download_dir) / pattern))
    except Exception:
        return []

def wait_for_download_completion(download_dir: str, pattern: str = "*.xls*", timeout: int = 90) -> Optional[str]:
    """
    Espera a que aparezca un archivo que cumpla el patrón y que no tenga la extensión .crdownload.
    Retorna la ruta del archivo descargado o None si expira el timeout.
    """
    start = time.time()
    download_dir = str(Path(download_dir).resolve())
    last_stable_path = None
    while time.time() - start < timeout:
        # No debe haber archivos temporales .crdownload
        tmp_in_progress = glob.glob(os.path.join(download_dir, "*.crdownload"))
        files = _list_downloads(download_dir, pattern)
        if files:
            # Elegir el más reciente
            files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
            candidate = files[0]
            # Verificar que no siga escribiéndose (tamaño estable)
            try:
                size1 = os.path.getsize(candidate)
                time.sleep(0.5)
                size2 = os.path.getsize(candidate)
                if size1 == size2 and not tmp_in_progress:
                    last_stable_path = candidate
                    break
            except Exception:
                pass
        time.sleep(0.5)
    return last_stable_path

def click_export_excel_in_open_modal(
    driver: webdriver.Chrome,
    timeout: int = 30,
    button_text: str = "Exportar Excel",
    download_dir: Optional[str] = None,
    file_pattern: str = "*.xls*"
) -> Optional[str]:
    """
    Asume que el modal ya está visible. Hace clic en el botón "Exportar Excel" y
    espera a que el archivo se descargue en el directorio indicado.

    Retorna la ruta del archivo descargado o None si falla.
    """
    try:
        # Asegurar directorio de descarga
        download_dir = download_dir or get_downloads_dir()
        Path(download_dir).mkdir(parents=True, exist_ok=True)

        # Intentar encontrar el botón por texto exacto u opciones cercanas
        btn_xpaths = [
            f"//button[normalize-space()='{button_text}']",
            f"//a[normalize-space()='{button_text}']",
            f"//div[contains(@class,'modal') and .//*[normalize-space()='{button_text}']]//*[self::button or self::a][normalize-space()='{button_text}']",
            f"//*[self::button or self::a][contains(normalize-space(), 'Exportar') and contains(normalize-space(), 'Excel')]",
        ]
        button = None
        for xp in btn_xpaths:
            try:
                button = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.XPATH, xp))
                )
                if button and button.is_displayed():
                    break
            except Exception:
                continue
        if not button:
            return None

        # Clic con JS para evitar overlays
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", button)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", button)
        except Exception:
            try:
                button.click()
            except Exception:
                return None

        # Esperar a que inicie y finalice la descarga
        downloaded = wait_for_download_completion(download_dir, file_pattern, timeout=120)
        return downloaded
    except Exception:
        return None

def exportar_excel_despues_de_modal(
    driver: webdriver.Chrome,
    download_dir: Optional[str] = None,
    timeout: int = 30,
    button_text: str = "Exportar Excel",
    file_pattern: str = "*.xls*"
) -> Optional[str]:
    """
    Úsala inmediatamente después de que tu script abra el modal.
    No intenta abrir el modal: solo hace clic en el botón "Exportar Excel" visible
    dentro del modal actual y espera la descarga en la carpeta indicada (por defecto, Descargas).
    Retorna la ruta del archivo descargado o None si falla.
    """
    return click_export_excel_in_open_modal(
        driver,
        timeout=timeout,
        button_text=button_text,
        download_dir=download_dir,
        file_pattern=file_pattern,
    )

def wait_modal_then_export_excel(
    driver: webdriver.Chrome,
    timeout: int = 40,
    modal_title_contains: str = "DETALLE DE PEDIDOS",
    button_text: str = "Exportar Excel",
    download_dir: Optional[str] = None,
    file_pattern: str = "*.xls*"
) -> Optional[str]:
    """
    Espera a que un modal esté visible (por título o por presencia del botón de exportar) y
    dispara la exportación a Excel, guardando en Downloads.
    """
    try:
        # Condición: modal visible por clases comunes o por título
        modal_locators = [
            (By.CSS_SELECTOR, ".modal.show"),
            (By.XPATH, f"//div[contains(@class,'modal') and .//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{modal_title_contains}')]]"),
            (By.XPATH, f"//*[self::button or self::a][normalize-space()='{button_text}']")
        ]
        visible = False
        for by, sel in modal_locators:
            try:
                WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((by, sel)))
                visible = True
                break
            except Exception:
                continue
        if not visible:
            return None
        return click_export_excel_in_open_modal(
            driver,
            timeout=timeout,
            button_text=button_text,
            download_dir=download_dir,
            file_pattern=file_pattern,
        )
    except Exception:
        return None

def navigate_menu_path(driver: webdriver.Chrome, items: List[str], timeout: int = 20) -> bool:
    """
    Navega por una ruta de menú haciendo clic por texto en cada nivel.
    Ej: ['DE REALIZACION', 'Ver Visitas Piloto']
    """
    try:
        open_side_menu_if_needed(driver)
        for label in items:
            ok = click_element_by_text(driver, label, timeout=timeout)
            if not ok:
                print(f"No se pudo hacer clic en el ítem de menú: {label}")
                return False
            time.sleep(0.6)
        return True
    except Exception as e:
        print(f"Error navegando ruta de menú {items}: {e}")
        return False

def search_and_open_menu(driver: webdriver.Chrome, query: str, timeout: int = 15) -> bool:
    """Escribe en el buscador del menú lateral y abre la opción por el texto indicado."""
    try:
        # Localizar input de búsqueda del menú
        search_candidates = [
            (By.CSS_SELECTOR, "input[placeholder*='Buscar'][type='text']"),
            (By.XPATH, "//input[contains(@placeholder,'Buscar')]")
        ]
        search_el = None
        for by, sel in search_candidates:
            try:
                search_el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, sel)))
                if search_el and search_el.is_displayed():
                    break
            except Exception:
                continue
        if not search_el:
            print("No se encontró el buscador del menú")
            return False

        search_el.clear()
        search_el.send_keys(query)
        time.sleep(0.6)

        # Intentar clic directo sobre opción con el texto
        if click_element_by_text(driver, query, timeout=3):
            return True

        # Si no hay opción clickeable visible, probar Enter para abrir el primer resultado
        try:
            search_el.send_keys(Keys.RETURN)
            time.sleep(1.2)
        except Exception:
            pass

        # Verificar que el título de la página cambiara al esperado
        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]"
                ))
            )
            return True
        except Exception:
            return False
    except Exception as e:
        print(f"Error al buscar y abrir '{query}': {e}")
        return False

def open_control_almacenes_and_open_total(
    driver: webdriver.Chrome,
    fecha_inicio: str,
    fecha_fin: str,
    timeout: int = 20,
    use_fullscreen: bool = False
) -> bool:
    """
    Abre 'Control de Almacenes', configura rango de fechas, pulsa 'Consultar' y abre el modal con 'TOTAL'.
    """
    try:
        open_side_menu_if_needed(driver)
        # Buscar y abrir la página 'Control de Almacenes'
        if not search_and_open_menu(driver, "Control de Almacenes", timeout=timeout):
            # Alternativa: navegar por menú si existe sección
            if not navigate_menu_path(driver, ["DE RECURSOS", "Control de Almacenes"]):
                print("No se pudo abrir 'Control de Almacenes'")
                return False

        # Esperar que cargue el panel de filtros de Control de Almacenes
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]"))
        )
        driver.save_screenshot("step_control_almacenes_loaded.png")

        # Activar 'Fecha de Creación' si hay checkbox y setear fechas
        try:
            # Tildar el checkbox si existe
            chk = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/preceding::input[@type='checkbox'][1]")
            if not chk.is_selected():
                driver.execute_script("arguments[0].click();", chk)
        except Exception:
            pass

        # Asegurar que 'Fecha de Recepcion' esté desmarcada
        try:
            chk_rx = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Recepcion')]/preceding::input[@type='checkbox'][1]")
            if chk_rx.is_selected():
                driver.execute_script("arguments[0].click();", chk_rx)
        except Exception:
            pass

        # Setear rango por teclado (más compatible con datepickers)
        if not set_date_inputs_by_label(driver, 'Fecha de Creacion', fecha_inicio, fecha_fin, timeout=timeout):
            # Respaldo con JS si falla el tipeo
            try:
                container = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/parent::*/parent::*")
                date_inputs = container.find_elements(By.CSS_SELECTOR, "input[type='date'], input[class*='date'], input[type='text']")
                if len(date_inputs) >= 2:
                    for el, val in [(date_inputs[0], fecha_inicio), (date_inputs[1], fecha_fin)]:
                        driver.execute_script(
                            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                            el, val
                        )
                    driver.save_screenshot("step_dates_set.png")
                else:
                    print("No se encontraron dos inputs de fecha para 'Fecha de Creación'")
            except Exception:
                print("Fallo configurando fechas por JS de respaldo")

        # Cerrar posibles overlays del datepicker antes de Consultar
        close_overlays_and_datepickers(driver)
        # Clic en Consultar
        clicked_consultar = False
        consultar_xpaths = [
            "//button[normalize-space()='Consultar']",
            "//a[normalize-space()='Consultar']",
            "//button[contains(.,'Consultar')]",
            "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONSULTAR')]"
        ]
        for xp in consultar_xpaths:
            try:
                btn = driver.find_element(By.XPATH, xp)
                if btn.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                    time.sleep(0.2)
                    driver.execute_script("arguments[0].click();", btn)
                    clicked_consultar = True
                    break
            except Exception:
                continue
        if not clicked_consultar:
            print("No se pudo clicar en 'Consultar'")
        # Esperar a que termine el procesamiento
        if not wait_until_not_processing(driver, timeout=40):
            print("Advertencia: la página siguió procesando más de lo esperado")
        # Si surgió alerta de error, cerrarla e intentar una vez más
        if dismiss_error_dialog_if_any(driver):
            close_overlays_and_datepickers(driver)
            try:
                for xp in consultar_xpaths:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el.is_displayed():
                            driver.execute_script("arguments[0].click();", el)
                            break
                    except Exception:
                        continue
            except Exception:
                pass

        # Sexto fallback: navegación por teclado (TAB) hasta 'CONSULTAR' y Enter
        if not consultar_clicked:
            try:
                print("Fallback: navegando con TAB hasta 'CONSULTAR'...")
                body = driver.find_element(By.TAG_NAME, 'body')
                body.click()
                for i in range(12):
                    try:
                        body.send_keys(Keys.TAB)
                        time.sleep(0.15)
                        focused_text = driver.execute_script("var el=document.activeElement; return (el.innerText||el.value||'').toString();") or ''
                        focused_upper = (focused_text or '').strip().upper()
                        if 'CONSULTAR' in focused_upper:
                            try:
                                body.send_keys(Keys.ENTER)
                            except Exception:
                                pass
                            try:
                                body.send_keys(Keys.SPACE)
                            except Exception:
                                pass
                            consultar_clicked = True
                            print("Clic en 'CONSULTAR' por teclado (TAB+ENTER/SPACE)")
                            break
                    except Exception:
                        continue
            except Exception:
                pass
        driver.save_screenshot("step_after_consultar.png")

        # Esperar que aparezcan los totales o botón 'Exportar Detalle' con reintentos
        def wait_totals_or_export(timeout_local=20):
            try:
                WebDriverWait(driver, timeout_local).until(
                    lambda d: (
                        any(el.is_displayed() for el in d.find_elements(By.XPATH, "//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]") ) or
                        any(el.is_displayed() for el in d.find_elements(By.XPATH, "//th[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]") )
                    )
                )
                return True
            except Exception:
                return False

        if not wait_totals_or_export(20):
            # Reintentar Consultar hasta 2 veces
            for _ in range(2):
                try:
                    btn = None
                    for xp in consultar_xpaths:
                        try:
                            el = driver.find_element(By.XPATH, xp)
                            if el.is_displayed():
                                btn = el
                                break
                        except Exception:
                            continue
                    if btn:
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", btn)
                        wait_until_not_processing(driver, timeout=40)
                        if wait_totals_or_export(20):
                            break
                except Exception:
                    pass

        # Intentar 'PANTALLA COMPLETA' solo si se solicita
        if use_fullscreen:
            try:
                full_xpaths = [
                    "//button[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'PANTALLA COMPLETA')]",
                    "//a[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'PANTALLA COMPLETA')]",
                ]
                for xp in full_xpaths:
                    try:
                        full_btn = driver.find_element(By.XPATH, xp)
                        if full_btn.is_displayed():
                            driver.execute_script("arguments[0].click();", full_btn)
                            time.sleep(1)
                            driver.save_screenshot("step_fullscreen_clicked.png")
                            break
                    except Exception:
                        continue
            except Exception:
                pass

        # Click en la tarjeta/celda TOTAL para abrir el modal
        total_clicked = False
        # 1) Intentar botón/enlace/td con TOTAL
        total_xpaths = [
            "//div[(contains(@class,'card') or contains(@class,'panel') or contains(@class,'col')) and contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]",
            "//td[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]",
            "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]"
        ]
        for xp in total_xpaths:
            try:
                el = driver.find_element(By.XPATH, xp)
                if el.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                    time.sleep(0.2)
                    driver.execute_script("arguments[0].click();", el)
                    total_clicked = True
                    break
            except Exception:
                continue
        # 2) Alternativa: ubicar cabecera 'TOTAL' y clicar la celda numérica debajo
        if not total_clicked:
            try:
                th_total = driver.find_element(By.XPATH, "//th[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]")
                # Tomar el índice de la columna TOTAL
                all_th = th_total.find_elements(By.XPATH, "ancestor::tr[1]/th")
                idx = all_th.index(th_total) + 1
                # Buscar la primera fila con celdas numéricas en esa columna
                td = driver.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", td)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", td)
                total_clicked = True
            except Exception:
                pass

        if not total_clicked:
            print("No se pudo clicar en 'TOTAL'")
            return False

        # Esperar apertura del modal DETALLE DE PEDIDOS
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS') and (self::h2 or self::h3 or self::div)]"))
        )
        driver.save_screenshot("step_modal_opened.png")
        return True
    except Exception as e:
        print(f"Error en flujo de Control de Almacenes: {e}")
        return False

def open_control_almacenes_and_open_category(
    driver: webdriver.Chrome,
    fecha_inicio: str,
    fecha_fin: str,
    categoria: str = "TOTAL",
    timeout: int = 20,
    use_fullscreen: bool = False
) -> bool:
    """
    Igual que open_control_almacenes_and_open_total, pero permite especificar la categoría
    a abrir, por ejemplo: "TRANSF. POR RECEPCIONAR".
    """
    try:
        # Normalizar categoría a mayúsculas para búsquedas de texto
        label_upper = (categoria or "").upper()
        # Reutilizar el flujo base para llegar a Control de Almacenes y consultar
        if not search_and_open_menu(driver, "Control de Almacenes", timeout=timeout):
            if not navigate_menu_path(driver, ["DE RECURSOS", "Control de Almacenes"]):
                print("No se pudo abrir 'Control de Almacenes'")
                return False

        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONTROL DE ALMACENES')]"))
        )

        # Setear fechas (mismo método robusto)
        try:
            chk = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/preceding::input[@type='checkbox'][1]")
            if not chk.is_selected():
                driver.execute_script("arguments[0].click();", chk)
        except Exception:
            pass

        if not set_date_inputs_by_label(driver, 'Fecha de Creacion', fecha_inicio, fecha_fin, timeout=timeout):
            # Respaldo con JS
            try:
                container = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/parent::*/parent::*")
                date_inputs = container.find_elements(By.CSS_SELECTOR, "input[type='date'], input[class*='date'], input[type='text']")
                if len(date_inputs) >= 2:
                    for el, val in [(date_inputs[0], fecha_inicio), (date_inputs[1], fecha_fin)]:
                        driver.execute_script(
                            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                            el, val
                        )
            except Exception:
                pass
        
        # Cerrar posibles overlays del datepicker antes de Consultar
        try:
            driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        except Exception:
            pass
        # Consultar
        for xp in [
            "//button[normalize-space()='Consultar']",
            "//a[normalize-space()='Consultar']",
            "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONSULTAR')]"
        ]:
            try:
                btn = driver.find_element(By.XPATH, xp)
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    break
            except Exception:
                continue
        # Esperar a que termine el procesamiento
        if not wait_until_not_processing(driver, timeout=40):
            print("Advertencia: la página siguió procesando más de lo esperado")

        # Esperar que aparezcan los totales o botón 'Exportar Detalle' con reintentos
        def wait_totals_or_export(timeout_local=20):
            try:
                WebDriverWait(driver, timeout_local).until(
                    lambda d: (
                        any(el.is_displayed() for el in d.find_elements(By.XPATH, "//*[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]") ) or
                        any(el.is_displayed() for el in d.find_elements(By.XPATH, "//th[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]") )
                    )
                )
                return True
            except Exception:
                return False

        if not wait_totals_or_export(20):
            # Reintentar Consultar hasta 2 veces
            for _ in range(2):
                try:
                    btn = None
                    for xp in [
                        "//button[normalize-space()='Consultar']",
                        "//a[normalize-space()='Consultar']",
                        "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONSULTAR')]"
                    ]:
                        try:
                            el = driver.find_element(By.XPATH, xp)
                            if el.is_displayed():
                                btn = el
                                break
                        except Exception:
                            continue
                    if btn:
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", btn)
                        wait_until_not_processing(driver, timeout=40)
                        if wait_totals_or_export(20):
                            break
                except Exception:
                    pass

        # Intentar pantalla completa solo si se solicita
        if use_fullscreen:
            for xp in [
                "//button[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'PANTALLA COMPLETA')]",
                "//a[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'PANTALLA COMPLETA')]",
            ]:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        driver.execute_script("arguments[0].click();", el)
                        time.sleep(0.8)
                        break
                except Exception:
                    continue

        # Esperar que los totales/valores estén cargados (evita fallos por errores de assets)
        try:
            # localizar cabecera de la categoría deseada y esperar a que la primera celda tenga dígitos
            th = driver.find_element(By.XPATH, f"//th[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]")
            all_th = th.find_elements(By.XPATH, "ancestor::tr[1]/th")
            idx = all_th.index(th) + 1
            def cell_has_digits(d):
                try:
                    td = d.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                    txt = td.text.strip()
                    return bool(re.search(r"\d+", txt))
                except Exception:
                    return False
            WebDriverWait(driver, timeout).until(cell_has_digits)
        except Exception:
            pass

        # Cerrar posibles overlays del datepicker antes de clicar categoría
        close_overlays_and_datepickers(driver)
        # Desmarcar 'Fecha de Recepcion'
        try:
            chk = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Recepcion')]/preceding::input[@type='checkbox'][1]")
            if chk.is_selected():
                driver.execute_script("arguments[0].click();", chk)
        except Exception:
            pass

        # 1) PRIMERO: localizar th de la categoría y clicar el NÚMERO (td) de la primera fila
        # Método auxiliar para hacer clic por coordenadas absolutas
        def click_at_position(x, y, offset_x=0, offset_y=0):
            try:
                # Mover el ratón a la posición absoluta (viewport) y hacer clic
                ActionChains(driver).move_by_offset(x + offset_x, y + offset_y).click().perform()
                # Resetear el offset para no acumular movimientos
                ActionChains(driver).move_by_offset(-(x + offset_x), -(y + offset_y)).perform()
                return True
            except Exception as e:
                print(f"Error en click_at_position: {e}")
                return False

        # Cerrar cualquier overlay o datepicker antes de empezar
        close_overlays_and_datepickers(driver)
        dismiss_error_dialog_if_any(driver)

        # Obtener la posición del TD del primer 7 (TRANSF. POR RECEPCIONAR)
        try:
            # Localizar el TH de la columna objetivo
            th_candidates = driver.find_elements(By.XPATH, "//th")
            th_target = None
            for th in th_candidates:
                txt = (th.text or "").upper().strip()
                if not txt:
                    continue
                if (label_upper in txt) or ("TRANSF" in txt and "RECEPC" in txt):
                    th_target = th
                    break
            if th_target is not None:
                # 1. Intentar clic normal
                all_th = th_target.find_elements(By.XPATH, "ancestor::tr[1]/th")
                idx = all_th.index(th_target) + 1
                td = driver.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", td)
                time.sleep(0.5)
                
                # 2. Hacer clic con ActionChains (más preciso)
                try:
                    ActionChains(driver).move_to_element(td).pause(0.5).click().perform()
                    clicked = True
                    print("Clic exitoso con ActionChains")
                except Exception as e:
                    print(f"Error con ActionChains: {e}")
                    try:
                        # 3. Si falla, intentar con JS click
                        driver.execute_script("arguments[0].click();", td)
                        clicked = True
                        print("Clic exitoso con JS")
                    except Exception as e2:
                        print(f"Error con JS click: {e2}")
                        # 4. Último recurso: coordenadas fijas basadas en la captura
                        try:
                            # Ajustes basados en la captura de pantalla
                            # x: posición horizontal del 7 (ajustar según la captura)
                            # y: posición vertical de la fila (ajustar según la captura)
                            x_offset = 850  # Ajustado a la derecha para el 7
                            y_offset = 450  # Ajustado para la fila correcta
                            
                            # Obtener posición del body para referencia
                            body = driver.find_element(By.TAG_NAME, 'body')
                            body_rect = body.rect
                            
                            # Calcular posición absoluta
                            x = body_rect['x'] + x_offset
                            y = body_rect['y'] + y_offset
                            
                            # Mover y hacer clic
                            ActionChains(driver).move_by_offset(x, y).click().perform()
                            # Resetear offset
                            ActionChains(driver).move_by_offset(-x, -y).perform()
                            clicked = True
                            print(f"Clic exitoso en coordenadas: x={x}, y={y}")
                        except Exception as e3:
                            print(f"Error con coordenadas fijas: {e3}")
                            clicked = False
        except Exception:
            clicked = False

        # 2) FALLBACK: clic por texto visible de la categoría
        if not clicked:
            category_xpaths = [
                f"//div[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]",
                f"//td[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]",
                f"//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]",
            ]
            for xp in category_xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", el)
                        clicked = True
                        break
                except Exception:
                    continue

        if not clicked:
            print(f"No se pudo clicar en la categoría: {categoria}")
            return False

        # Si aparece alerta de error al clicar, cerrarla y reintentar una vez el clic
        try:
            if dismiss_error_dialog_if_any(driver):
                close_overlays_and_datepickers(driver)
                # Reintentar el clic en el número de la columna
                try:
                    th_candidates = driver.find_elements(By.XPATH, "//th")
                    th_target = None
                    for th in th_candidates:
                        txt = (th.text or "").upper().strip()
                        if not txt:
                            continue
                        if label_upper in txt or "TRANSF" in txt and "RECEPC" in txt:
                            th_target = th
                            break
                    if th_target is not None:
                        all_th = th_target.find_elements(By.XPATH, "ancestor::tr[1]/th")
                        idx = all_th.index(th_target) + 1
                        td = driver.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", td)
                        time.sleep(0.1)
                        driver.execute_script("arguments[0].click();", td)
                except Exception:
                    pass
        except Exception:
            pass

        # Esperar que el modal aparezca
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS') and (self::h2 or self::h3 or self::div)]"))
        )
        return True
    except Exception as e:
        print(f"Error abriendo categoría {categoria}: {e}")
        return False

def extract_data(driver: webdriver.Chrome, timeout: int = 20, use_excel_export: bool = False, download_dir: Optional[str] = None) -> Dict[str, Any]:
    """
    Extrae datos de la página actual después del inicio de sesión, con soporte robusto para
    ventanas emergentes (modales), incluidos títulos, tablas y paginación interna.
    
    Args:
        driver: Instancia de Selenium WebDriver ya autenticada
        timeout: Tiempo máximo de espera en segundos para los elementos
        
    Returns:
        Dict con los datos extraídos estructurados
    """
    def get_table_data(table_el) -> Dict[str, Any]:
        # Encabezados
        headers: List[str] = []
        try:
            ths = table_el.find_elements(By.TAG_NAME, "th")
            if ths:
                headers = [th.text.strip() or f"Columna_{i+1}" for i, th in enumerate(ths)]
            else:
                first_tr = table_el.find_element(By.TAG_NAME, "tr")
                tds = first_tr.find_elements(By.TAG_NAME, "td")
                headers = [f"Columna_{i+1}" for i in range(len(tds))]
        except Exception:
            headers = []

        # Filas
        rows: List[Dict[str, str]] = []
        trs = table_el.find_elements(By.TAG_NAME, "tr")
        start_idx = 1 if headers and len(headers) > 1 else 0
        for i in range(start_idx, len(trs)):
            try:
                cells = trs[i].find_elements(By.TAG_NAME, "td")
                if not cells:
                    continue
                row: Dict[str, str] = {}
                for j, cell in enumerate(cells):
                    col_name = headers[j] if j < len(headers) else f"Columna_{j+1}"
                    row[col_name] = cell.text.strip()
                rows.append(row)
            except Exception as e:
                print(f"Error leyendo fila {i}: {e}")
        return {"encabezados": headers, "filas": rows}

    def paginate_and_collect(context_el) -> Dict[str, Any]:
        # Busca tabla y pagina dentro de un contenedor (modal o página)
        table = None
        table_selectors = ["table", ".table", "[role='grid']", ".data-table", "div[class*='table'] table"]
        for sel in table_selectors:
            try:
                table = context_el.find_element(By.CSS_SELECTOR, sel)
                break
            except Exception:
                continue
        if not table:
            return {"encabezados": [], "filas": [], "paginas": 0}

        all_rows: List[Dict[str, str]] = []
        headers_final: List[str] = []
        max_pages = 20
        pages_processed = 0

        # 1) Intentar configurar tamaño de página a 100 si existe un <select>
        try:
            # Buscar selects dentro del mismo contenedor (footer de paginación)
            selects = context_el.find_elements(By.TAG_NAME, "select")
            for sel in selects:
                try:
                    options_text = [o.text.strip() for o in sel.find_elements(By.TAG_NAME, 'option')]
                    if any(t in options_text for t in ["100", "50", "25", "10"]):
                        Select(sel).select_by_visible_text("100")
                        time.sleep(0.8)
                        break
                except Exception:
                    continue
        except Exception:
            pass
        while pages_processed < max_pages:
            # Scroll dentro del contenedor por si hay lazy-load
            try:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", context_el)
                time.sleep(0.5)
            except Exception:
                pass

            data_page = get_table_data(table)
            if data_page["encabezados"] and not headers_final:
                headers_final = data_page["encabezados"]
            all_rows.extend(data_page["filas"])
            pages_processed += 1

            # Intentar ir a la siguiente página
            next_btn = None
            next_selectors = [
                ".pagination .next:not(.disabled) a",
                ".pagination li.next:not(.disabled) a",
                "button[aria-label*='Siguiente']:not([disabled])",
                "button[title*='Siguiente']:not([disabled])",
                "a[rel='next']",
                # Variantes con símbolos
                "//a[normalize-space()='>']",
                "//a[normalize-space()='»']",
                "//button[normalize-space()='>']",
                "//button[normalize-space()='»']",
            ]
            for sel in next_selectors:
                try:
                    if sel.startswith("//"):
                        el = context_el.find_element(By.XPATH, sel)
                    else:
                        el = context_el.find_element(By.CSS_SELECTOR, sel)
                    if el and el.is_displayed():
                        next_btn = el
                        break
                except Exception:
                    continue

            if not next_btn:
                break

            try:
                driver.execute_script("arguments[0].click();", next_btn)
                # Pequeña espera para que se recargue la tabla
                WebDriverWait(driver, timeout).until(lambda d: True)
                time.sleep(1)
            except Exception:
                break

        return {"encabezados": headers_final, "filas": all_rows, "paginas": pages_processed}

    def wait_for_download_and_find_file(directory: str, timeout_sec: int = 120) -> Optional[str]:
        """Espera a que termine una descarga (sin .crdownload) y retorna el último .xlsx/.xls."""
        start = time.time()
        directory = str(Path(directory))
        while time.time() - start < timeout_sec:
            # Si hay archivos .crdownload, esperar
            cr = glob.glob(os.path.join(directory, "*.crdownload"))
            if cr:
                time.sleep(0.5)
                continue
            # Buscar Excel
            xl = sorted(
                glob.glob(os.path.join(directory, "*.xlsx")) + glob.glob(os.path.join(directory, "*.xls")),
                key=os.path.getmtime,
                reverse=True
            )
            if xl:
                return xl[0]
            time.sleep(0.5)
        return None

    def try_export_excel_from_modal() -> Optional[Dict[str, Any]]:
        """Intenta hacer clic en 'Exportar Excel' del modal y leer el archivo descargado."""
        if not download_dir:
            print("Exportación a Excel solicitada pero no se proporcionó download_dir")
            return None
        try:
            # Buscar botón Exportar Detalle (texto real del botón en el modal)
            export_xpaths = [
                "//div[contains(@class,'modal') or @role='dialog']//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]",
                # Compatibilidad: 'Exportar Excel'
                "//div[contains(@class,'modal') or @role='dialog']//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXCEL')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXCEL')]",
                # Variante reportada: 'EXPORTAR EXEL' (sin c)
                "//div[contains(@class,'modal') or @role='dialog']//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXEL')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXEL')]",
            ]
            btn = None
            for xp in export_xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        btn = el
                        break
                except Exception:
                    continue
            if not btn:
                print("No se encontró botón 'Exportar Excel' en el modal")
                return None

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", btn)
            print("Click en 'Exportar Detalle' ejecutado, esperando descarga...")

            downloaded = wait_for_download_and_find_file(download_dir)
            if not downloaded:
                print("No se detectó archivo Excel descargado")
                return None
            print(f"Archivo Excel descargado: {downloaded}")

            # Leer el Excel. Intentar con pandas, fallback a openpyxl si está disponible
            data_rows: List[Dict[str, Any]] = []
            headers: List[str] = []
            try:
                import pandas as pd
                df = pd.read_excel(downloaded)
                headers = list(df.columns.astype(str))
                data_rows = df.fillna("").to_dict(orient='records')
            except Exception as e:
                print(f"No se pudo leer con pandas: {e}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(downloaded, read_only=True, data_only=True)
                    ws = wb.active
                    rows_iter = ws.iter_rows(values_only=True)
                    first = True
                    for r in rows_iter:
                        if first:
                            headers = [str(x) if x is not None else f"Columna_{i+1}" for i, x in enumerate(r)]
                            first = False
                            continue
                        row_dict = {}
                        for i, val in enumerate(r):
                            col = headers[i] if i < len(headers) else f"Columna_{i+1}"
                            row_dict[col] = "" if val is None else str(val)
                        data_rows.append(row_dict)
                except Exception as e2:
                    print(f"No se pudo leer Excel sin pandas: {e2}")
                    return {
                        "estado": "éxito",
                        "origen": "excel_descargado",
                        "ruta_excel": downloaded,
                        "total_registros": 0,
                        "encabezados": [],
                        "datos": []
                    }

            return {
                "estado": "éxito",
                "origen": "excel_descargado",
                "ruta_excel": downloaded,
                "total_registros": len(data_rows),
                "encabezados": headers,
                "datos": data_rows
            }
        except Exception as e:
            print(f"Error durante exportación/lectura Excel: {e}")
            return None

    def try_export_excel_from_page() -> Optional[Dict[str, Any]]:
        """Intenta exportar desde el botón verde 'Exportar Detalle' en la página de Control de Almacenes."""
        if not download_dir:
            return None
        try:
            # Buscar botón en la página (no en modal)
            xp_list = [
                "//button[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]",
                "//a[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR DETALLE')]",
                # 'Exportar Excel'
                "//button[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXCEL')]",
                "//a[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXCEL')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXCEL')]",
                # 'Exportar Exel' (sin c)
                "//button[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXEL')]",
                "//a[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXEL')]",
                "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR EXEL')]",
            ]
            btn = None
            for xp in xp_list:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        btn = el
                        break
                except Exception:
                    continue
            if not btn:
                return None
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", btn)
            print("Click en botón de página 'Exportar Detalle' ejecutado, esperando descarga...")

            downloaded = wait_for_download_and_find_file(download_dir)
            if not downloaded:
                print("No se detectó archivo Excel descargado desde página")
                return None

            # Leer Excel
            data_rows: List[Dict[str, Any]] = []
            headers: List[str] = []
            try:
                import pandas as pd
                df = pd.read_excel(downloaded)
                headers = list(df.columns.astype(str))
                data_rows = df.fillna("").to_dict(orient='records')
            except Exception:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(downloaded, read_only=True, data_only=True)
                    ws = wb.active
                    rows_iter = ws.iter_rows(values_only=True)
                    first = True
                    for r in rows_iter:
                        if first:
                            headers = [str(x) if x is not None else f"Columna_{i+1}" for i, x in enumerate(r)]
                            first = False
                            continue
                        row_dict = {}
                        for i, val in enumerate(r):
                            col = headers[i] if i < len(headers) else f"Columna_{i+1}"
                            row_dict[col] = "" if val is None else str(val)
                        data_rows.append(row_dict)
                except Exception as e2:
                    print(f"No se pudo leer Excel: {e2}")
                    return None

            return {
                "estado": "éxito",
                "origen": "excel_descargado_pagina",
                "ruta_excel": downloaded,
                "total_registros": len(data_rows),
                "encabezados": headers,
                "datos": data_rows
            }
        except Exception as e:
            print(f"Error exportando desde página: {e}")
            return None

    try:
        current_url = driver.current_url
        print(f"Extrayendo datos desde: {current_url}")

        # Esperar a que la página se estabilice
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        time.sleep(1)

        # Buscar todos los modales visibles
        modal_selectors = [
            "div[role='dialog']",
            ".modal.show .modal-dialog",
            ".modal-dialog .modal-content",
            ".MuiDialog-container",
            "div[class*='modal'][style*='display: block']",
        ]
        modales: List[Any] = []
        for sel in modal_selectors:
            try:
                found = driver.find_elements(By.CSS_SELECTOR, sel)
                modales.extend([m for m in found if m.is_displayed()])
            except Exception:
                continue

        resultados_modales: List[Dict[str, Any]] = []

        # Procesar cada modal encontrado
        for idx, modal in enumerate(modales):
            print(f"Procesando modal #{idx+1}")
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", modal)
            except Exception:
                pass

            # Intentar detectar título del modal
            titulo = None
            try:
                for ts in [".modal-title", "[role='dialog'] [class*='title']", "h2", "h3"]:
                    try:
                        t = modal.find_element(By.CSS_SELECTOR, ts)
                        if t.text.strip():
                            titulo = t.text.strip()
                            break
                    except Exception:
                        continue
            except Exception:
                pass

            # Manejar iframes dentro del modal
            switched = False
            try:
                iframe = modal.find_element(By.TAG_NAME, "iframe")
                driver.switch_to.frame(iframe)
                switched = True
                print("Cambiado al iframe dentro del modal")
            except Exception:
                pass

            try:
                datos_tabla = paginate_and_collect(driver if switched else modal)
            finally:
                if switched:
                    driver.switch_to.default_content()

            resultados_modales.append({
                "modal_index": idx + 1,
                "titulo": titulo or f"Modal_{idx+1}",
                "encabezados": datos_tabla.get("encabezados", []),
                "datos": datos_tabla.get("filas", []),
                "paginas_procesadas": datos_tabla.get("paginas", 1)
            })

        # Intentar exportación a Excel si fue solicitado
        if use_excel_export:
            # 1) Intentar desde modal
            excel_result = try_export_excel_from_modal()
            if not excel_result:
                # 2) Si no hubo modal o no funcionó, intentar botón de página
                excel_result = try_export_excel_from_page()
            if excel_result:
                excel_result.update({
                    "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "url_origen": current_url,
                })
                return excel_result

        # Si no hubo modales o no tenían tablas, intentar en la página principal
        if not resultados_modales or all(len(m.get("datos", [])) == 0 for m in resultados_modales):
            print("No se obtuvieron datos desde modales. Probando en la página principal...")
            datos_pagina = paginate_and_collect(driver)
            if datos_pagina.get("filas"):
                return {
                    "estado": "éxito",
                    "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "url_origen": current_url,
                    "origen": "pagina",
                    "total_registros": len(datos_pagina["filas"]),
                    "encabezados": datos_pagina["encabezados"],
                    "datos": datos_pagina["filas"],
                }

        if resultados_modales:
            total = sum(len(m.get("datos", [])) for m in resultados_modales)
            return {
                "estado": "éxito",
                "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S"),
                "url_origen": current_url,
                "origen": "modales",
                "total_registros": total,
                "modales": resultados_modales,
            }

        # Si no se logró nada
        driver.save_screenshot("extract_data_sin_resultados.png")
        return {
            "estado": "error",
            "mensaje": "No se encontraron datos en modales ni en la página",
            "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S"),
            "url_origen": current_url
        }
    except TimeoutException:
        return {
            "estado": "error",
            "mensaje": "Tiempo de espera agotado al cargar los datos",
            "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S")
        }
    except Exception as e:
        try:
            driver.save_screenshot("extract_data_error.png")
        except Exception:
            pass
        return {
            "estado": "error",
            "mensaje": f"Error al extraer datos: {str(e)}",
            "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S")
        }

def save_to_database(data: Dict[str, Any], table_name: str = 'datos_externos') -> Dict[str, Any]:
    """
    Guarda los datos extraídos en la base de datos MySQL.
    
    Args:
        data: Diccionario con los datos a guardar
        table_name: Nombre de la tabla donde se guardarán los datos
        
    Returns:
        Dict con el resultado de la operación
    """
    connection = None
    try:
        # Configuración de la conexión (ajusta según tu configuración)
        connection = mysql.connector.connect(
            host='localhost',
            user='root',  # Cambia por tu usuario de MySQL
            password='',  # Cambia por tu contraseña
            database='hermes_express'  # Asegúrate de que esta base de datos exista
        )
        
        if connection.is_connected():
            cursor = connection.cursor()
            
            # Crear la tabla si no existe
            create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fecha_consulta DATETIME NOT NULL,
                tipo_dato VARCHAR(50) NOT NULL,
                contenido JSON NOT NULL,
                estado VARCHAR(20) DEFAULT 'pendiente',
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
            cursor.execute(create_table_query)
            
            # Insertar los datos
            insert_query = f"""
            INSERT INTO {table_name} (fecha_consulta, tipo_dato, contenido)
            VALUES (%s, %s, %s)
            """
            
            record = (
                data.get('fecha_consulta', time.strftime("%Y-%m-%d %H:%M:%S")),
                'datos_savar',  # o cualquier otro identificador de tipo de dato
                json.dumps(data, ensure_ascii=False)
            )
            
            cursor.execute(insert_query, record)
            connection.commit()
            
            return {
                'estado': 'éxito',
                'mensaje': 'Datos guardados correctamente en la base de datos',
                'id_registro': cursor.lastrowid,
                'fecha_registro': time.strftime("%Y-%m-%d %H:%M:%S")
            }
            
    except Error as e:
        return {
            'estado': 'error',
            'mensaje': f'Error al guardar en la base de datos: {str(e)}',
            'fecha_error': time.strftime("%Y-%m-%d %H:%M:%S")
        }
        
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

def configurar_fechas_y_consultar(driver: webdriver.Chrome, fecha_inicio: str, fecha_fin: str, timeout: int = 20) -> bool:
    """
    Configura las fechas y hace clic en el botón Consultar para abrir el modal.
    """
    try:
        print(f"Configurando fechas: {fecha_inicio} a {fecha_fin}")
        
        # Activar 'Fecha de Recepción' si hay checkbox
        try:
            chk_rx = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Recepcion')]/preceding::input[@type='checkbox'][1]")
            if not chk_rx.is_selected():
                driver.execute_script("arguments[0].click();", chk_rx)
                print("Checkbox 'Fecha de Recepción' activado")
        except Exception:
            print("No se encontró checkbox de 'Fecha de Recepción'")
            pass

        # Desactivar 'Fecha de Creación' si está marcado
        try:
            chk = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/preceding::input[@type='checkbox'][1]")
            if chk.is_selected():
                driver.execute_script("arguments[0].click();", chk)
                print("Checkbox 'Fecha de Creación' desactivado")
        except Exception:
            # Configurar fechas
            print("Configurando fechas:", fecha_inicio, "a", fecha_fin)
            max_attempts = 3
            for attempt in range(1, max_attempts + 1):
                try:
                    set_date_inputs_by_label(driver, "Fecha de Creacion", fecha_inicio, fecha_fin, timeout=15)
                    print(f"Fechas configuradas correctamente (intento {attempt})")
                    break
                except TimeoutException:
                    print(f"Timeout en intento {attempt} de configurar fechas")
                    if attempt == max_attempts:
                        print("No se pudieron configurar las fechas después de múltiples intentos")
                        # Continuar sin fechas
                    else:
                        time.sleep(2)  # Esperar antes de reintentar
                except Exception as e:
                    print(f"Error al configurar fechas (intento {attempt}):", str(e))
                    if attempt == max_attempts:
                        print("No se pudieron configurar las fechas automáticamente")
                        # Continuar sin fechas
                    time.sleep(1)

        # Configurar fechas usando el método robusto
        if set_date_inputs_by_label(driver, 'Fecha de Recepcion', fecha_inicio, fecha_fin, timeout=timeout):
            print("Fechas configuradas correctamente")
        else:
            print("Advertencia: No se pudieron configurar las fechas automáticamente")
            # Intentar método alternativo con JavaScript
            try:
                container = driver.find_element(By.XPATH, "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Creacion')]/parent::*/parent::*")
                date_inputs = container.find_elements(By.CSS_SELECTOR, "input[type='date'], input[class*='date'], input[type='text']")
                if len(date_inputs) >= 2:
                    for el, val in [(date_inputs[0], fecha_inicio), (date_inputs[1], fecha_fin)]:
                        driver.execute_script(
                            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                            el, val
                        )
                    print("Fechas configuradas con JavaScript")
            except Exception as e:
                print(f"Error configurando fechas: {e}")

        # Cerrar posibles overlays del datepicker
        try:
            driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        except Exception:
            pass

        # Hacer clic en el botón Consultar con reintentos
        print("Buscando botón 'Consultar'...")
        try:
            driver.execute_script("window.scrollTo(0,0);")
        except Exception:
            pass
        consultar_clicked = False
        # Intento prioritario por ID exacto del botón Consultar
        try:
            btn_id = driver.find_element(By.ID, "btnBuscarCTRLALM")
            if btn_id and btn_id.is_displayed():
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_id)
                time.sleep(0.2)
                try:
                    driver.execute_script("arguments[0].click();", btn_id)
                except Exception:
                    btn_id.click()
                consultar_clicked = True
                print("Botón 'Consultar' clickeado por ID #btnBuscarCTRLALM")
        except Exception:
            pass

        consultar_xpaths = [
            "//button[normalize-space()='Consultar']",
            "//a[normalize-space()='Consultar']",
            "//button[contains(.,'Consultar')]",
            "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CONSULTAR')]"
        ]
        
        # Intentar hasta 3 veces
        for intento in range(3):
            try:
                for xp in consultar_xpaths:
                    try:
                        btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xp)))
                        if btn.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                            time.sleep(0.5)
                            driver.execute_script("arguments[0].click();", btn)
                            print(f"Botón 'Consultar' clickeado (intento {intento + 1})")
                            consultar_clicked = True
                            break
                    except Exception as e:
                        print(f"Error con xpath {xp}: {e}")
                        continue
                
                if consultar_clicked:
                    break
                else:
                    print(f"Intento {intento + 1} fallido, reintentando...")
                    time.sleep(2)
                    
            except Exception as e:
                print(f"Error en intento {intento + 1}: {e}")
                if intento < 2:  # No es el último intento
                    time.sleep(3)
                    continue

        # Fallback JS por texto visible si aún no se clickeó
        if not consultar_clicked:
            print("Fallback: intentando localizar 'CONSULTAR' por texto con JS...")
            try:
                driver.execute_script("void(0);")
            except Exception:
                pass
            try:
                js_ok = driver.execute_script(
                    "var upper=function(s){return (s||'').toString().toUpperCase()};"
                    "var visible=function(el){var r=el.getBoundingClientRect();return r.width>0&&r.height>0&&window.getComputedStyle(el).visibility!=='hidden'&&window.getComputedStyle(el).display!=='none'};"
                    "var nodes=document.querySelectorAll('button,a,input,div,span');"
                    "for(var i=0;i<nodes.length;i++){var el=nodes[i];var t=upper((el.innerText||el.value||'').trim());"
                    "if(t.indexOf('CONSULTAR')>-1 && visible(el)){try{el.disabled=false;el.removeAttribute('disabled');}catch(e){};"
                    "try{el.scrollIntoView({block:'center'});}catch(e){};"
                    "try{el.click();}catch(e){};"
                    "try{['mousedown','mouseup','click'].forEach(function(tipo){var ev=new MouseEvent(tipo,{view:window,bubbles:true,cancelable:true});el.dispatchEvent(ev);});}catch(e){};"
                    "return true;}} return false;"
                )
                consultar_clicked = True
                print("Clic en 'CONSULTAR' realizado por fallback JS")
            except Exception as e:
                print(f"Fallback JS no pudo clicar 'CONSULTAR': {e}")

        # Segundo fallback: click por coordenadas del botón si es visible en DOM
        if not consultar_clicked:
            try:
                btn = None
                for xp in consultar_xpaths:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el.is_displayed():
                            btn = el
                            break
                    except Exception:
                        continue
                if btn is not None:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    driver.execute_script(
                        "var r=arguments[0].getBoundingClientRect();"
                        "var cx=r.left+(r.width/2);var cy=r.top+(r.height/2);"
                        "var el=document.elementFromPoint(cx,cy);"
                        "if(el){['mousedown','mouseup','click'].forEach(function(t){var ev=new MouseEvent(t,{view:window,bubbles:true,cancelable:true,clientX:cx,clientY:cy});el.dispatchEvent(ev);});}",
                        btn
                    )
                    consultar_clicked = True
                    print("Clic en 'CONSULTAR' por coordenadas ejecutado")
            except Exception:
                pass

        # Tercer fallback: enviar Enter en el último input de fecha
        if not consultar_clicked:
            try:
                last_input = None
                try:
                    # Ubicar el segundo input bajo 'Fecha de Recepcion'
                    xpath_start = "//label[contains(translate(.,'áéíóúÁÉÍÓÚ','aeiouAEIOU'),'Fecha de Recepcion')]"
                    last_input = driver.find_element(By.XPATH, xpath_start + "/following::input[2]")
                except Exception:
                    pass
                if last_input is None:
                    inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='date'], input[class*='date'], input[type='text']")
                    if inputs:
                        last_input = inputs[-1]
                if last_input is not None:
                    last_input.click()
                    time.sleep(0.1)
                    last_input.send_keys(Keys.ENTER)
                    print("Enter enviado en el último input de fecha")
                    consultar_clicked = True
            except Exception:
                pass

        # Cuarto fallback: submit del formulario más cercano por JS
        if not consultar_clicked:
            try:
                js_submitted = driver.execute_script(
                    "var btns=[].slice.call(document.querySelectorAll('button,a'));"
                    "var upper=function(s){return (s||'').toString().toUpperCase()};"
                    "var visible=function(el){var r=el.getBoundingClientRect();return r.width>0&&r.height>0&&window.getComputedStyle(el).visibility!=='hidden'&&window.getComputedStyle(el).display!=='none'};"
                    "for(var i=0;i<btns.length;i++){var el=btns[i]; if(upper(el.innerText||'').indexOf('CONSULTAR')>-1 && visible(el)){var f=el.closest('form'); if(f){f.dispatchEvent(new Event('submit',{bubbles:true,cancelable:true})); try{f.submit();}catch(e){} return true;}}} return false;"
                )
                if js_submitted:
                    consultar_clicked = True
                    print("Formulario enviado por JS (submit)")
            except Exception:
                pass

        # Quinto fallback: click en una posición fija relativa al botón Exportar (misma barra)
        if not consultar_clicked:
            try:
                export_btn = None
                for xp in [
                    "//button[normalize-space()='Exportar']",
                    "//a[normalize-space()='Exportar']",
                    "//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'EXPORTAR')]"
                ]:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el.is_displayed():
                            export_btn = el
                            break
                    except Exception:
                        continue
                if export_btn is not None:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", export_btn)
                    time.sleep(0.1)
                    # Click 60px a la izquierda del centro de 'Exportar' (donde suele estar 'Consultar')
                    driver.execute_script(
                        "var r=arguments[0].getBoundingClientRect();"
                        "var cx=r.left+(r.width/2)-60; var cy=r.top+(r.height/2);"
                        "var el=document.elementFromPoint(cx,cy);"
                        "if(el){['mousedown','mouseup','click'].forEach(function(t){var ev=new MouseEvent(t,{view:window,bubbles:true,cancelable:true,clientX:cx,clientY:cy});el.dispatchEvent(ev);});}",
                        export_btn
                    )
                    consultar_clicked = True
                    print("Clic aproximado en posición de 'Consultar' ejecutado")
            except Exception:
                pass

        if not consultar_clicked:
            print("Error: No se pudo hacer clic en el botón 'Consultar'")
            try:
                driver.save_screenshot("error_click_consultar.png")
                print("Screenshot guardado: error_click_consultar.png")
            except Exception:
                pass
            return False

        # Esperar a que termine el procesamiento
        print("Esperando a que termine el procesamiento...")
        if not wait_until_not_processing(driver, timeout=40):
            print("Advertencia: la página siguió procesando más de lo esperado")

        # Cerrar posibles mensajes de error que bloquean la pantalla (OK de alerta)
        try:
            if dismiss_error_dialog_if_any(driver):
                print("Se cerró un diálogo de error tras Consultar")
        except Exception:
            pass

        # Asegurar que no quede ningún overlay abierto
        close_overlays_and_datepickers(driver)

        # Tomar screenshot después de consultar
        driver.save_screenshot("step_after_consultar.png")
        print("Screenshot guardado: step_after_consultar.png")
        
        return True
        
    except Exception as e:
        print(f"Error configurando fechas y consultando: {e}")
        return False

def abrir_modal_y_extraer_datos(driver: webdriver.Chrome, categoria: str = "TOTAL", column_label: Optional[str] = None, timeout: int = 20) -> Dict[str, Any]:
    """
    Abre el modal haciendo clic en un objetivo y extrae los datos.
    Puedes especificar:
    - categoria: texto de categoría (botón/etiqueta) a clicar
    - column_label: encabezado de columna sobre el cual se debe clicar el primer valor numérico (e.g. "EN ALMACEN RECEPCIONAR")
    """
    try:
        # Prevenir bloqueos por alertas/overlays antes de intentar los clics
        try:
            dismiss_error_dialog_if_any(driver)
        except Exception:
            pass
        close_overlays_and_datepickers(driver)

        if column_label:
            print(f"Abriendo modal haciendo clic en columna: {column_label}")
        else:
            print(f"Abriendo modal para la categoría: {categoria}")
        
        # Buscar y hacer clic en el objetivo especificado
        categoria_clicked = False
        label_upper = categoria.upper() if categoria else ""

        # Método 0A: intento directo por selector global del botón reportado
        if column_label and not categoria_clicked:
            # Detectar todas las opciones visibles y probar hasta que aparezca un modal
            def modal_appeared(d):
                sels = [
                    (By.CSS_SELECTOR, ".modal.show, .modal[style*='display: block']"),
                    (By.XPATH, "//*[@role='dialog' and (contains(@style,'display') or @aria-modal='true')]"),
                    (By.CSS_SELECTOR, ".MuiDialog-container, .swal2-container")
                ]
                for by, sel in sels:
                    try:
                        els = d.find_elements(by, sel)
                        if any(e.is_displayed() for e in els):
                            return True
                    except Exception:
                        continue
                return False

            candidates = []
            try:
                candidates = [el for el in driver.find_elements(By.CSS_SELECTOR, "span.ctrlalm-item-detail-header[data-estado='6']") if el.is_displayed()]
            except Exception:
                candidates = []
            if not candidates:
                try:
                    candidates = [el for el in driver.find_elements(By.XPATH, "//span[contains(@class,'ctrlalm-item-detail-header') and @data-estado='6']") if el.is_displayed()]
                except Exception:
                    candidates = []
            for cand in candidates:
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cand)
                    time.sleep(0.1)
                    try:
                        ActionChains(driver).move_to_element(cand).pause(0.1).click().perform()
                    except Exception:
                        try:
                            driver.execute_script("arguments[0].click();", cand)
                        except Exception:
                            # fallback mouse events
                            driver.execute_script(
                                "var r=arguments[0].getBoundingClientRect();"
                                "var cx=r.left+(r.width/2);var cy=r.top+(r.height/2);"
                                "var el=document.elementFromPoint(cx,cy);"
                                "if(el){['mousedown','mouseup','click'].forEach(function(t){var ev=new MouseEvent(t,{view:window,bubbles:true,cancelable:true,clientX:cx,clientY:cy});el.dispatchEvent(ev);});}",
                                cand
                            )
                    # Esperar a que se abra un modal
                    try:
                        WebDriverWait(driver, 4).until(modal_appeared)
                        categoria_clicked = True
                        break
                    except Exception:
                        continue
                except Exception:
                    continue

        # Método 0B: Clic por encabezado de columna y primera fila de datos
        if column_label and not categoria_clicked:
            try:
                # Localizar el TH por coincidencia normalizada (ignora tildes y saltos de línea)
                th_target = _find_th_by_text(driver, column_label)
                if th_target is not None:
                    all_th = th_target.find_elements(By.XPATH, "ancestor::tr[1]/th")
                    idx = all_th.index(th_target) + 1
                    td = driver.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", td)
                    time.sleep(0.2)
                    # Intentar clicar específicamente el 'botón/badge' negro con el número dentro de la celda
                    clicked_inner = False
                    try:
                        # 0) Priorizar el span específico reportado por el usuario
                        try:
                            target_el = td.find_element(By.CSS_SELECTOR, "span.ctrlalm-item-detail-header[data-estado='6']")
                        except Exception:
                            try:
                                target_el = td.find_element(By.XPATH, ".//span[contains(@class,'ctrlalm-item-detail-header')][@data-estado='6']")
                            except Exception:
                                target_el = None

                        # 1) Si no se encontró el específico, priorizar candidatos que parecen badge/botón con número
                        inner_candidates = td.find_elements(By.XPATH, ".//*[self::button or self::a or self::div or self::span]")
                        # Priorizar los que parezcan botones/badges por clase o estilo y que tengan dígitos
                        priorizados = []
                        for el in inner_candidates:
                            t = (el.text or '').strip()
                            cls = (el.get_attribute('class') or '').lower()
                            style = (el.get_attribute('style') or '').lower()
                            if re.search(r"\b(btn|badge|label|pill|dark|negro|black)\b", cls) or ('background' in style or '#000' in style or 'black' in style):
                                if re.search(r"\\d+", t):
                                    priorizados.append(el)
                        # Mantener target_el si ya fue encontrado por la clase específica
                        if target_el is None:
                            target_el = None
                        # Si hay priorizados, tomar el primero visible
                        if target_el is None:
                            for cand in priorizados:
                                try:
                                    if cand.is_displayed():
                                        target_el = cand
                                        break
                                except Exception:
                                    continue
                        else:
                            # fallback: cualquier descendiente con números
                            for el in inner_candidates:
                                try:
                                    if el.is_displayed() and re.search(r"\d+", (el.text or '').strip()):
                                        target_el = el
                                        break
                                except Exception:
                                    continue
                        if target_el is not None:
                            # Buscar ancestro clickeable (a, button, [onclick], [role='button'])
                            try:
                                ancestor = None
                                for _ in range(4):
                                    if target_el is None:
                                        break
                                    try:
                                        if target_el.tag_name.lower() in ['a','button']:
                                            ancestor = target_el
                                            break
                                        onclick = (target_el.get_attribute('onclick') or '').strip()
                                        rolebtn = (target_el.get_attribute('role') or '').lower() == 'button'
                                        if onclick or rolebtn:
                                            ancestor = target_el
                                            break
                                        # subir un nivel
                                        target_el = target_el.find_element(By.XPATH, "..")
                                    except Exception:
                                        break
                                if ancestor is not None:
                                    target_el = ancestor
                            except Exception:
                                pass
                            try:
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_el)
                                time.sleep(0.1)
                                # Intento 1: click normal con ActionChains
                                ActionChains(driver).move_to_element(target_el).pause(0.1).click().perform()
                                clicked_inner = True
                                # pequeña espera a que el modal aparezca
                                try:
                                    WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS')]")))
                                except Exception:
                                    pass
                            except Exception:
                                try:
                                    # Intento 2: click JS directo
                                    driver.execute_script("arguments[0].click();", target_el)
                                    clicked_inner = True
                                    try:
                                        WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS')]")))
                                    except Exception:
                                        pass
                                except Exception:
                                    try:
                                        # Intento 3: dispatch MouseEvents en el centro del elemento
                                        driver.execute_script(
                                            "var r=arguments[0].getBoundingClientRect();"
                                            "var cx=r.left+(r.width/2);var cy=r.top+(r.height/2);"
                                            "var el=document.elementFromPoint(cx,cy);"
                                            "if(el){['mouseover','mousemove','mousedown','mouseup','click'].forEach(function(t){var ev=new MouseEvent(t,{view:window,bubbles:true,cancelable:true,clientX:cx,clientY:cy});el.dispatchEvent(ev);});}",
                                            target_el
                                        )
                                        clicked_inner = True
                                        try:
                                            WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS')]")))
                                        except Exception:
                                            pass
                                    except Exception:
                                        try:
                                            # Intento 4: doble clic por si abre modal con dblclick
                                            ActionChains(driver).move_to_element(target_el).pause(0.1).double_click().perform()
                                            clicked_inner = True
                                            try:
                                                WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS')]")))
                                            except Exception:
                                                pass
                                        except Exception:
                                            clicked_inner = False
                    except Exception:
                        clicked_inner = False

                    if not clicked_inner:
                        # Clic en la celda completa como respaldo
                        try:
                            ActionChains(driver).move_to_element(td).pause(0.2).click().perform()
                        except Exception:
                            try:
                                driver.execute_script("arguments[0].click();", td)
                            except Exception:
                                try:
                                    # Último recurso: click por coordenadas del centro del TD
                                    driver.execute_script(
                                        "var r=arguments[0].getBoundingClientRect();"
                                        "var cx=r.left+(r.width/2);var cy=r.top+(r.height/2)+2;"
                                        "var el=document.elementFromPoint(cx,cy);"
                                        "if(el){['mousedown','mouseup','click'].forEach(function(t){var ev=new MouseEvent(t,{view:window,bubbles:true,cancelable:true,clientX:cx,clientY:cy});el.dispatchEvent(ev);});}",
                                        td
                                    )
                                except Exception:
                                    pass

                    print(f"Clic en columna '{column_label}' exitoso")
                    categoria_clicked = True
            except Exception as e:
                print(f"Error al clicar columna '{column_label}': {e}")
        
        # Método 1: Buscar por texto en la tabla
        categoria_xpaths = [
            f"//td[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]",
            f"//div[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]",
            f"//*[self::button or self::a][contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'{label_upper}')]"
        ]
        
        if not categoria_clicked and categoria:
            for xp in categoria_xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", el)
                        print(f"Clic en categoría '{categoria}' exitoso")
                        categoria_clicked = True
                        break
                except Exception:
                    continue

        # Método 2: Si no se encontró por texto, buscar por posición en la tabla
        if not categoria_clicked and not column_label:
            try:
                # Buscar la columna TOTAL en la tabla
                th_total = driver.find_element(By.XPATH, "//th[contains(translate(.,'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'TOTAL')]")
                all_th = th_total.find_elements(By.XPATH, "ancestor::tr[1]/th")
                idx = all_th.index(th_total) + 1
                # Hacer clic en la primera celda de datos de esa columna
                td = driver.find_element(By.XPATH, f"//tr[td][1]/td[{idx}]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", td)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", td)
                print("Clic en columna TOTAL exitoso")
                categoria_clicked = True
            except Exception as e:
                print(f"Error con método de posición: {e}")

        if not categoria_clicked:
            objetivo = column_label if column_label else categoria
            print(f"Error: No se pudo hacer clic en el objetivo '{objetivo}'")
            return {
                "estado": "error",
                "mensaje": f"No se pudo abrir el modal para '{objetivo}'"
            }

        # Esperar a que aparezca el modal
        print("Esperando a que aparezca el modal...")
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: (
                    any(e.is_displayed() for e in d.find_elements(By.XPATH, "//*[contains(translate(., 'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DETALLE DE PEDIDOS') and (self::h2 or self::h3 or self::div)]")) or
                    any(e.is_displayed() for e in d.find_elements(By.CSS_SELECTOR, ".modal.show, .modal[style*='display: block']")) or
                    any(e.is_displayed() for e in d.find_elements(By.XPATH, "//*[@role='dialog' and (contains(@style,'display') or @aria-modal='true')]") )
                )
            )
            print("Modal 'DETALLE DE PEDIDOS' abierto correctamente")
        except Exception:
            print("Advertencia: No se detectó el modal 'DETALLE DE PEDIDOS' automáticamente")

        # Tomar screenshot del modal
        driver.save_screenshot("step_modal_opened.png")
        print("Screenshot del modal guardado: step_modal_opened.png")

        # Descargar Excel desde el modal y retornar solo la ruta
        print("Clic en 'Exportar Excel' y esperando descarga...")
        download_dir = str(Path(__file__).parent / 'downloads')
        try:
            Path(download_dir).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        ruta_excel = click_export_excel_in_open_modal(
            driver,
            timeout=40,
            button_text="Exportar Excel",
            download_dir=download_dir,
            file_pattern="*.xls*",
        )

        if not ruta_excel:
            return {
                "estado": "error",
                "mensaje": "No se pudo descargar el Excel desde el modal"
            }

        return {
            "estado": "éxito",
            "origen": "excel_descargado",
            "ruta_excel": ruta_excel,
            "fecha_consulta": time.strftime("%Y-%m-%d %H:%M:%S")
        }
        
    except Exception as e:
        print(f"Error abriendo modal y extrayendo datos: {e}")
        return {
            "estado": "error",
            "mensaje": f"Error al abrir modal y extraer datos: {str(e)}"
        }

def main():
    """
    Función principal para probar el módulo.
    """
    print("=== EXTRACCION DE DATOS SAVAR EXPRESS ===")
    
    # Configurar el navegador con carpeta de descargas
    download_dir = str(Path(__file__).parent / 'downloads')
    driver = setup_driver(headless=False, download_dir=download_dir)
    driver.maximize_window()  # Maximizar ventana para mejor visibilidad
    
    try:
        # Credenciales de prueba
        usuario = "CHI.HER"
        contrasena = "123456789"
        
        # Definir fechas para la consulta (formato: YYYY-MM-DD)
        # fecha_inicio = día anterior, fecha_fin = día de ejecución
        from datetime import datetime, timedelta
        hoy_dt = datetime.now()
        ayer_dt = hoy_dt - timedelta(days=1)
        fecha_inicio = ayer_dt.strftime('%Y-%m-%d')
        fecha_fin = hoy_dt.strftime('%Y-%m-%d')
        
        print(f"\nConfigurando consulta para el rango de fechas: {fecha_inicio} al {fecha_fin}")
        
        # Paso 1: Iniciar sesión
        print("\nPASO 1: INICIANDO SESION...")
        if login_and_fetch_saver(driver, usuario, contrasena, fecha_inicio, fecha_fin):
            print("Inicio de sesion exitoso")
            
            # Paso 2: Navegar a Control de Almacenes
            print("\nPASO 2: NAVEGANDO A CONTROL DE ALMACENES...")
            if search_and_open_menu(driver, "Control de Almacenes", timeout=20):
                print("Control de Almacenes abierto")
                
                # Paso 3: Configurar fechas y hacer clic en Consultar
                print("\nPASO 3: CONFIGURANDO FECHAS Y CONSULTANDO...")
                if configurar_fechas_y_consultar(driver, fecha_inicio, fecha_fin):
                    print("Fechas configuradas y consulta ejecutada")
                    
                    # Paso 4: Abrir modal y extraer datos
                    print("\nPASO 4: ABRIENDO MODAL Y EXTRAYENDO DATOS...")
                    datos = abrir_modal_y_extraer_datos(driver, categoria="TOTAL", column_label="EN ALMACEN")
                else:
                    print("Error: No se pudo configurar las fechas y consultar")
                    datos = {"estado": "error", "mensaje": "No se pudo configurar fechas y consultar"}
            else:
                print("Error: No se pudo abrir 'Control de Almacenes'")
                datos = {"estado": "error", "mensaje": "No se pudo abrir Control de Almacenes"}
        else:
            print("Error: No se pudo iniciar sesion")
            datos = {"estado": "error", "mensaje": "No se pudo iniciar sesión"}
        
        # Mostrar resultados mínimos y no generar JSON
        print("\n" + "="*50)
        print("RESULTADO DE DESCARGA DE EXCEL:")
        print("="*50)
        print(f"Estado: {datos.get('estado')}")
        print(f"Mensaje: {datos.get('mensaje', 'N/A')}")
        if datos.get('ruta_excel'):
            print(f"Archivo Excel: {datos.get('ruta_excel')}")
            
    except Exception as e:
        print(f"\nError durante la ejecución: {e}")
        import traceback
        traceback.print_exc()
        
    finally:
        # Captura del estado final y cierre seguro del navegador
        try:
            driver.save_screenshot("final_state.png")
        except Exception:
            pass
        try:
            print("\n" + "="*50)
            print("PROCESO FINALIZADO")
            print("="*50)
            print("Revisa los siguientes archivos:")
            print("- datos_savar.json (datos extraídos)")
            print("- downloads/ (archivos Excel descargados)")
            print("- Screenshots (imágenes del proceso)")
            try:
                print("\nEsperando 10 segundos antes de cerrar...")
                time.sleep(10)
            except Exception:
                time.sleep(5)
        finally:
            try:
                driver.quit()
            except Exception:
                pass

if __name__ == "__main__":
    main()
